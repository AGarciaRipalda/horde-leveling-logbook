import openpyxl
import json
import re

# File paths
EXCEL_FILE = "WOW TBC Leveling Route_ 60-70.xlsx"
TEMPLATE_FILE = "index_complete.html" # Alliance
OUTPUT_FILE = "horde-leveling-logbook/index.html" # Result

def extract_horde_data():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # 1. Introduction
    intro_sheet = wb['Introduction']
    intro_data = {
        "general_info": [],
        "xp_table": []
    }
    
    # Extract General Info (rows 2-8 based on previous check)
    for row in intro_sheet.iter_rows(min_row=2, max_row=8, values_only=True):
        if row[0]: # Column A
            intro_data["general_info"].append(row[0])
            
    # Extract XP Table (rows 10-20 based on structure)
    # XP table headers at row 9? Let's check or assume standard
    # Based on index_complete.html, XP table is standard.
    # I'll hardcode it as fallback or check if sheet has it.
    # Given the complexity, I'll reuse the XP table from the Alliance index as it's standard TBC XP.
    xp_table = [
        { "level": 61, "xp_needed": 494000 },
        { "level": 62, "xp_needed": 574700 },
        { "level": 63, "xp_needed": 614400 },
        { "level": 64, "xp_needed": 650300 },
        { "level": 65, "xp_needed": 682300 },
        { "level": 66, "xp_needed": 710200 },
        { "level": 67, "xp_needed": 734100 },
        { "level": 68, "xp_needed": 753700 },
        { "level": 69, "xp_needed": 768900 },
        { "level": 70, "xp_needed": 779700 }
    ]
    intro_data["xp_table"] = xp_table

    # 2. Route
    route_sheet = wb['The Route']
    route_data = []
    
    # Iterate from row 2 (assuming row 1 is header)
    task_id = 1
    for row in route_sheet.iter_rows(min_row=2, values_only=True):
        # Stop if row is empty (all None)
        if all(cell is None for cell in row):
            continue
            
        # Map columns: A=Type, B=Name, C=Notes, D=Rewards
        task_type = row[0]
        task_name = row[1]
        task_notes = row[2]
        task_rewards = row[3]
        
        # Skip header rows if encountered again (e.g. "Quest", "Task", ...)
        if task_type == "Type" or task_name == "Task":
            continue

        if task_type:
            task_type = str(task_type).lower().strip()
        else:
            task_type = "other" # Default if missing
            
        route_item = {
            "id": task_id,
            "type": task_type,
            "name": task_name,
            "notes": task_notes,
            "rewards": task_rewards
        }
        route_data.append(route_item)
        task_id += 1
        
    return {
        "introduction": intro_data,
        "route": route_data
    }

def update_index_file(data):
    # Read template (Alliance index)
    with open(TEMPLATE_FILE, "r", encoding="utf-8") as f:
        content = f.read()
        
    # Prepare JSON string
    json_str = json.dumps(data, indent=2)
    
    # Replace DATA const safely using brace counting
    start_marker = "const DATA = {"
    start_index = content.find(start_marker)
    
    if start_index == -1:
        print("Error: Could not find 'const DATA = {'")
        return

    # Find the end of the object by counting braces
    open_braces = 0
    end_index = -1
    in_string = False
    escape = False
    
    # Start scanning from the opening brace of DATA
    scan_start = start_index + len("const DATA = ")
    
    for i, char in enumerate(content[scan_start:], start=scan_start):
        if char == '"' and not escape:
            in_string = not in_string
        
        if not in_string:
            if char == '{':
                open_braces += 1
            elif char == '}':
                open_braces -= 1
                if open_braces == 0:
                    end_index = i + 1 # Include the closing brace
                    break
        
        if char == '\\' and not escape:
            escape = True
        else:
            escape = False
            
    if end_index == -1:
        print("Error: Could not find matching closing brace for DATA")
        return
        
    # Check if there is a semicolon after
    if end_index < len(content) and content[end_index] == ';':
        end_index += 1
        
    print(f"Replacing DATA block from index {start_index} to {end_index}")
    
    new_content = content[:start_index] + f"const DATA = {json_str};" + content[end_index:]
    
    # Write to Output file
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(new_content)
        
    print(f"Successfully updated {OUTPUT_FILE} with {len(data['route'])} tasks.")

if __name__ == "__main__":
    try:
        data = extract_horde_data()
        update_index_file(data)
    except Exception as e:
        print(f"Error: {e}")
